
import os
from openpyxl import load_workbook
import smtplib
from cryptography.fernet import Fernet

#Read the Excel with the Data
filenm = int(input("""Ingrese la institucion a enviar emails: \n
                1 - Area 1\n
                2 - Hernandarias \n
                3 - UNE\n"""))
if not filenm in [1, 2, 3]:
    raise Exception("Institucion no reconocida")
if filenm == 1:
    wb = load_workbook(filename = "Salas de Área 1.xlsx")
    sheets = [name for name in wb.sheetnames if not ("rea 1" in name)]
    place = "el Colegio Nacional Atanasio Riera Area 1"
    times = ["7:30AM", "8:00AM", "12:00PM"]

elif filenm == 2:
    wb = load_workbook(filename = "Salas de Hernandarias.xlsx")
    sheets = [name for name in wb.sheetnames if not ("Tacur" in name)]
    place = "el colegio Tacuru Pucu"
    times = ["12:30PM", "1:00PM", "5:00PM"]
elif filenm == 3:
    wb = load_workbook(filename = "Salas UNE.xlsx")
    sheets = [name for name in wb.sheetnames if not ("UNE" in name)]
    place = "la Universidad Nacional del Este Campus KM7 - Facultad de Ciencias Economicas"
    times = ["7:30AM", "8:00AM", "12:00PM"]


with open("E:\\.Trashes\\hola\\mnk.bin", "rb") as f:
    F = Fernet(f.read())

if True:
    with open("email.key", "rb") as f:
        correo = F.decrypt(f.read()).decode()
    with open("mykey.key", "rb") as f:
        passs = F.decrypt(f.read()).decode()

#authentification steps
s = smtplib.SMTP(host = "smtp.gmail.com", port = 587)
s.ehlo()
s.starttls()
s.ehlo()
s.login(correo, passs)

for name in sheets:
    sheet = wb[name]


    senders = []
    c = 2
    if not ("orreo" in sheet['D1'].value.lower()):
        raise Exception("Advertencia: La columna D puede no ser la de correos")
    while sheet['D{}'.format(c)].value :
        senders.append(sheet['D{}'.format(c)].value)
        c+=1

    try:

        from email.mime.multipart import MIMEMultipart as mimu
        from email.mime.text import MIMEText as mimtxt
        from email.mime.base import MIMEBase as mimbase
        from email.header import Header
        from email import encoders

        msg = mimu()
        msg['From'] = "Team Mathura"
        #msg['To'] = "Voluntariado"
        msg['Subject'] = 'Confirmacion Seleccion Mathura 2019'
        att = "Colores\\" + name.lower() + ".png"
        with open(att, 'rb') as f:
            base = mimbase('image', 'png', filename = 'attach.png')
            base.add_header('Content-Dispotition', 'attachment; filename="attach.png"')
            #base.add_header('X-Attachment-Id', '0')
            #base.add_header('Content-ID', '<0>')
            base.set_payload(f.read())
            encoders.encode_base64(base)
            msg.attach(base)

        with open("msgs.html") as f:
            body = f.read()
        myname = (name[:-1] + "a") if name[-1] == "o" else name
        myname = "Oro" if myname.lower() == "ora" else myname
        body = body.format(times[0], place, times[1], times[2], name)
        teext = mimtxt(body, 'html', 'utf-8')

        msg.attach(teext)


        s.sendmail(correo, senders, msg.as_string())


        #here the email has been already sent


    except FileNotFoundError as e:
        print(e)
        print("Hubo un error con el archivo decodificador")
s.quit()


del passs, correo, s
'''
if __name__ == "__main__":
    main()
'''
